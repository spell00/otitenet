
# /home/simon/otitenet/otitenet/app/services/inference_results_service.py

from __future__ import annotations

import copy
import glob
import os
import re
import unicodedata
from typing import Any, Dict, Iterable, List, Optional, Tuple

import numpy as np
import pandas as pd
import streamlit as st
from sklearn.metrics import accuracy_score, matthews_corrcoef, roc_auc_score


# -------------------------------------------------
# Formatting helpers
# -------------------------------------------------

def fmt_metric(value, digits: int = 4) -> str:
    """Format a scalar metric for Streamlit."""
    try:
        if value is None:
            return "—"
        value = float(value)
        if np.isnan(value):
            return "—"
        return f"{value:.{digits}f}"
    except Exception:
        return "—"


def fmt_confidence(value, digits: int = 4) -> str:
    """Format confidence/probability values safely."""
    try:
        if value is None:
            return "—"
        value = float(value)
        if np.isnan(value):
            return "—"
        return f"{value:.{digits}f}"
    except Exception:
        return "—"


# -------------------------------------------------
# Label normalization / correctness
# -------------------------------------------------

def _normalize_label(label) -> str:
    """
    Normalize labels for robust comparison.

    Handles common spelling/case variations:
    - NotNormal / not normal / abnormal
    - Normal / healthy
    """
    if label is None:
        return ""

    text = str(label).strip()
    if text == "":
        return ""

    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "", text)

    aliases = {
        "normal": "normal",
        "healthy": "normal",
        "notnormal": "notnormal",
        "nothealthy": "notnormal",
        "abnormal": "notnormal",
        "otitis": "notnormal",
        "otite": "notnormal",
        "malade": "notnormal",
        "infected": "notnormal",
        "infection": "notnormal",
    }

    return aliases.get(text, text)


def labels_match(pred_label, true_label):
    """
    Return True/False/None for prediction correctness.

    Returns None when ground truth is unavailable or unknown.
    """
    pred = _normalize_label(pred_label)
    truth = _normalize_label(true_label)

    if not truth or truth in {"unknown", "na", "nan", "none"}:
        return None

    if not pred or pred in {"unknown", "na", "nan", "none"}:
        return None

    return pred == truth


# -------------------------------------------------
# Ground truth loading from Excel
# -------------------------------------------------

GROUND_TRUTH_WORKBOOK_CANDIDATES = [
    os.environ.get("OTITENET_GT_XLSX"),
    "LOG_Entrainement_IA.xlsx",
    "LOG_Entrainement_IA(1).xlsx",
    "./LOG_Entrainement_IA.xlsx",
    "./LOG_Entrainement_IA(1).xlsx",
    "data/LOG_Entrainement_IA.xlsx",
    "data/LOG_Entrainement_IA(1).xlsx",
    "data/datasets/LOG_Entrainement_IA.xlsx",
    "data/datasets/LOG_Entrainement_IA(1).xlsx",
    "/home/simon/otitenet/LOG_Entrainement_IA.xlsx",
    "/home/simon/otitenet/LOG_Entrainement_IA(1).xlsx",
]


def _norm_filename_for_gt(path_or_name) -> str:
    """Normalize filenames for ground-truth lookup."""
    if path_or_name is None:
        return ""

    name = os.path.basename(str(path_or_name).strip())
    name = unicodedata.normalize("NFC", name)
    return name.lower()


def _candidate_gt_workbooks(inference_dir=None) -> List[str]:
    """
    Return possible workbook paths.

    The expected workbook contains:
      sheet: classifications
      column A: path
      column B: class
    """
    candidates = []

    for path in GROUND_TRUTH_WORKBOOK_CANDIDATES:
        if path:
            candidates.append(path)

    if inference_dir:
        inference_dir = str(inference_dir)
        candidates.extend(
            [
                os.path.join(inference_dir, "LOG_Entrainement_IA.xlsx"),
                os.path.join(inference_dir, "LOG_Entrainement_IA(1).xlsx"),
                os.path.join(os.path.dirname(inference_dir), "LOG_Entrainement_IA.xlsx"),
                os.path.join(os.path.dirname(inference_dir), "LOG_Entrainement_IA(1).xlsx"),
                os.path.join(os.path.dirname(os.path.dirname(inference_dir)), "LOG_Entrainement_IA.xlsx"),
                os.path.join(os.path.dirname(os.path.dirname(inference_dir)), "LOG_Entrainement_IA(1).xlsx"),
            ]
        )

    # Bounded fallback search for copied workbooks.
    for root in [".", "data", "data/datasets", "/home/simon/otitenet"]:
        try:
            candidates.extend(
                glob.glob(
                    os.path.join(root, "**", "LOG_Entrainement_IA*.xlsx"),
                    recursive=True,
                )
            )
        except Exception:
            pass

    out = []
    seen = set()

    for path in candidates:
        if not path:
            continue

        path = os.path.abspath(os.path.expanduser(str(path)))

        if path not in seen:
            out.append(path)
            seen.add(path)

    return out


def _read_classifications_with_openpyxl(xlsx_path: str) -> Dict[str, str]:
    """
    Read ground truth using openpyxl.

    Expected:
      sheet name: classifications
      column A: path
      column B: class
    """
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    if "classifications" not in wb.sheetnames:
        raise ValueError(f"Workbook has no sheet named 'classifications'. Found: {wb.sheetnames}")

    ws = wb["classifications"]
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        return {}

    header = [str(x).strip().lower() if x is not None else "" for x in rows[0]]

    # Prefer named columns but fall back to user-specified positions.
    try:
        path_idx = header.index("path")
    except ValueError:
        path_idx = 0

    try:
        class_idx = header.index("class")
    except ValueError:
        class_idx = 1

    mapping = {}

    for row in rows[1:]:
        if not row or len(row) <= max(path_idx, class_idx):
            continue

        path_val = row[path_idx]
        class_val = row[class_idx]

        if path_val is None or class_val is None:
            continue

        key = _norm_filename_for_gt(path_val)
        label = str(class_val).strip()

        if key and label:
            mapping[key] = label

    return mapping


def _read_classifications_with_pandas(xlsx_path: str) -> Dict[str, str]:
    """
    Fallback reader using pandas.

    Important: user confirmed column B of sheet 'classifications' is the class.
    """
    df = pd.read_excel(xlsx_path, sheet_name="classifications")

    if df.shape[1] < 2:
        return {}

    path_col = df.columns[0]
    class_col = df.columns[1]

    mapping = {}

    for _, row in df.iterrows():
        path_val = row.get(path_col)
        class_val = row.get(class_col)

        if pd.isna(path_val) or pd.isna(class_val):
            continue

        key = _norm_filename_for_gt(path_val)
        label = str(class_val).strip()

        if key and label:
            mapping[key] = label

    return mapping


@st.cache_data(show_spinner=False)
def _load_ground_truth_map_cached(inference_dir=None):
    """
    Load ground-truth map from LOG_Entrainement_IA*.xlsx.

    Returns:
      mapping, workbook_path, error
    """
    errors = []

    for xlsx_path in _candidate_gt_workbooks(inference_dir):
        if not os.path.exists(xlsx_path):
            continue

        try:
            mapping = _read_classifications_with_openpyxl(xlsx_path)
            if mapping:
                return mapping, xlsx_path, None
        except Exception as exc:
            errors.append(f"{xlsx_path}: openpyxl failed: {exc}")

        try:
            mapping = _read_classifications_with_pandas(xlsx_path)
            if mapping:
                return mapping, xlsx_path, None
        except Exception as exc:
            errors.append(f"{xlsx_path}: pandas failed: {exc}")

    return {}, None, " | ".join(errors[-5:])


def inference_ground_truth(filename, inference_dir="data/datasets/inference"):
    """
    Return ground truth class for an inference image.

    The source of truth is:
      Excel workbook: LOG_Entrainement_IA*.xlsx
      Sheet: classifications
      Column A: path
      Column B: class

    Matching is by basename, so both of these match:
      data/datasets/inference/00001-D.jpeg
      00001-D.jpeg
    """
    mapping, _xlsx_path, _error = _load_ground_truth_map_cached(inference_dir)

    key = _norm_filename_for_gt(filename)

    if key in mapping:
        return mapping[key]

    # Extension-insensitive fallback.
    key_no_ext = os.path.splitext(key)[0]

    for stored_key, label in mapping.items():
        if os.path.splitext(stored_key)[0] == key_no_ext:
            return label

    return "Unknown"


def ground_truth_source(inference_dir="data/datasets/inference"):
    """Small diagnostic helper for UI/debugging."""
    mapping, xlsx_path, error = _load_ground_truth_map_cached(inference_dir)
    return {
        "workbook": xlsx_path,
        "n_labels": len(mapping),
        "error": error,
    }


# -------------------------------------------------
# Inference metrics
# -------------------------------------------------

def compute_inference_metrics(df: pd.DataFrame) -> Dict[str, Any]:
    """
    Compute ACC/MCC/AUC from a dataframe with:
      Ground Truth, Prediction, Correct, Confidence

    AUC is only computed when binary labels and numeric confidence are available.
    """
    if df is None or len(df) == 0:
        return {"ACC": np.nan, "MCC": np.nan, "AUC": np.nan, "N": 0}

    work = df.copy()

    if "Ground Truth" not in work.columns or "Prediction" not in work.columns:
        return {"ACC": np.nan, "MCC": np.nan, "AUC": np.nan, "N": len(work)}

    work["_truth_norm"] = work["Ground Truth"].apply(_normalize_label)
    work["_pred_norm"] = work["Prediction"].apply(_normalize_label)

    valid = work[
        (work["_truth_norm"].notna())
        & (work["_pred_norm"].notna())
        & (~work["_truth_norm"].isin(["", "unknown", "na", "nan", "none"]))
        & (~work["_pred_norm"].isin(["", "unknown", "na", "nan", "none"]))
    ].copy()

    if valid.empty:
        return {"ACC": np.nan, "MCC": np.nan, "AUC": np.nan, "N": 0}

    y_true = valid["_truth_norm"].values
    y_pred = valid["_pred_norm"].values

    try:
        acc = float(accuracy_score(y_true, y_pred))
    except Exception:
        acc = np.nan

    try:
        mcc = float(matthews_corrcoef(y_true, y_pred))
    except Exception:
        mcc = np.nan

    auc = np.nan
    try:
        # Binary AUC only. Treat notnormal as positive if present.
        classes = sorted(set(y_true))
        if len(classes) == 2 and "Confidence" in valid.columns:
            positive = "notnormal" if "notnormal" in classes else classes[-1]
            y_bin = np.array([1 if y == positive else 0 for y in y_true])
            conf = pd.to_numeric(valid["Confidence"], errors="coerce").values

            if not np.isnan(conf).all() and len(np.unique(y_bin)) == 2:
                # If confidence is for the predicted class, flip it when the predicted
                # class is the negative class. This is approximate but useful for dashboarding.
                pred_is_pos = np.array([1 if p == positive else 0 for p in y_pred])
                pos_score = np.where(pred_is_pos == 1, conf, 1.0 - conf)
                auc = float(roc_auc_score(y_bin, pos_score))
    except Exception:
        auc = np.nan

    out = {
        "ACC": acc,
        "MCC": mcc,
        "AUC": auc,
        "N": int(len(valid)),
    }

    for label in sorted(set(y_true) | set(y_pred)):
        true_pos = y_true == label
        pred_pos = y_pred == label
        tp = int(np.sum(true_pos & pred_pos))
        fp = int(np.sum(~true_pos & pred_pos))
        fn = int(np.sum(true_pos & ~pred_pos))
        support = int(np.sum(true_pos))

        precision = tp / (tp + fp) if (tp + fp) else np.nan
        recall = tp / (tp + fn) if (tp + fn) else np.nan
        f1 = (
            2 * precision * recall / (precision + recall)
            if np.isfinite(precision) and np.isfinite(recall) and (precision + recall)
            else np.nan
        )
        display_label = str(label)
        out[f"F1 {display_label}"] = f1
        out[f"Recall {display_label}"] = recall
        out[f"Precision {display_label}"] = precision
        out[f"Support {display_label}"] = support

    return out


# -------------------------------------------------
# Model-row to args helper
# -------------------------------------------------

def _row_get(row: Dict[str, Any], *keys, default=None):
    for key in keys:
        if key in row and row.get(key) is not None:
            return row.get(key)
    return default


def args_from_inference_row(base_args, row_dict: Dict[str, Any]):
    """
    Build a model args namespace from a leaderboard/inference row.

    Used by inference_results.py and ensemble.py.
    """
    try:
        args = copy.copy(base_args)
    except Exception:
        args = base_args

    row = dict(row_dict or {})

    model_id = _row_get(row, "Model ID", "model_id", "id")
    if model_id is not None:
        try:
            args.model_id = int(model_id)
        except Exception:
            args.model_id = model_id

    mapping = {
        "model_name": ("Model Name", "model_name", "Model"),
        "new_size": ("NSize", "nsize", "Size"),
        "fgsm": ("FGSM", "fgsm"),
        "prototypes_to_use": ("Prototypes", "prototypes", "prototype"),
        "n_positives": ("NPos", "npos", "n_positives"),
        "n_negatives": ("NNeg", "nneg", "n_negatives"),
        "dloss": ("DLoss", "dloss"),
        "dist_fct": ("Dist_Fct", "dist_fct", "Dist"),
        "classif_loss": ("Classif_Loss", "classif_loss", "Loss"),
        "n_calibration": ("N_Calibration", "n_calibration", "N_Cal"),
        "normalize": ("Normalize", "normalize"),
        "n_neighbors": ("N_Neighbors", "n_neighbors"),
        "path": ("Path", "path"),
        "task": ("Task", "task"),
        "log_path": ("Log Path", "log_path"),
        "prototype_strategy": ("Proto_Strat", "prototype_strategy"),
        "prototype_components": ("Proto_Comp", "prototype_components"),
    }

    for attr, keys in mapping.items():
        value = _row_get(row, *keys)
        if value is None:
            continue

        if attr in {"new_size", "n_neighbors", "n_positives", "n_negatives", "prototype_components"}:
            try:
                value = int(float(value))
            except Exception:
                pass

        setattr(args, attr, value)

    # Some rows only have log_path, not path. Leave path unchanged unless row has path.
    if not hasattr(args, "bs"):
        args.bs = 32
    if not hasattr(args, "groupkfold"):
        args.groupkfold = 1
    if not hasattr(args, "random_recs"):
        args.random_recs = 0
    if not hasattr(args, "task"):
        args.task = "otite"

    return args


# -------------------------------------------------
# Analysis result normalization
# -------------------------------------------------

def normalize_analysis_result(result):
    """
    Normalize run_analysis_on_file(...) return shapes.

    Supported shapes:
      (pred_label, confidence)
      (pred_label, confidence, complete_log_path)
      (pred_label, confidence, complete_log_path, existing)
      (pred_label, confidence, complete_log_path, existing, gradcam_path)
    """
    if isinstance(result, dict):
        pred_label = result.get("pred_label", result.get("Prediction", result.get("prediction")))
        confidence = result.get("confidence", result.get("Confidence"))
        complete_log_path = result.get("complete_log_path", result.get("log_path", result.get("Log Path")))
        existing = result.get("existing", result.get("Existing"))
        gradcam_path = result.get("gradcam_path", result.get("Grad-CAM Path"))
        return pred_label, confidence, complete_log_path, existing, gradcam_path

    if not isinstance(result, (tuple, list)):
        raise ValueError(f"Unsupported analysis result type: {type(result)}")

    if len(result) < 2:
        raise ValueError(f"Analysis result has too few values: {len(result)}")

    pred_label = result[0]
    confidence = result[1]
    complete_log_path = result[2] if len(result) > 2 else None
    existing = result[3] if len(result) > 3 else None
    gradcam_path = result[4] if len(result) > 4 else None

    return pred_label, confidence, complete_log_path, existing, gradcam_path


# -------------------------------------------------
# Grad-CAM discovery
# -------------------------------------------------

def _strip_extension(filename: str) -> str:
    return os.path.splitext(os.path.basename(str(filename)))[0]


def find_inference_gradcam_images(log_path, filename, n_layers: int = 4) -> List[str]:
    """
    Find Grad-CAM images generated during inference/new analysis.

    Expected current layout:
      <log_path>/<base>/<base>_grad_cam_all_classes_layerX.png

    Also searches legacy/fallback patterns under log_path.
    """
    if not log_path or not filename:
        return []

    base = _strip_extension(filename)
    log_path = str(log_path).rstrip("/")

    roots = [
        log_path,
        os.path.join(log_path, base),
        os.path.join(log_path, "queries"),
        os.path.join(log_path, "queries", base),
    ]

    patterns = []

    for root in roots:
        patterns.extend(
            [
                os.path.join(root, f"{base}_grad_cam_all_classes_layer*.png"),
                os.path.join(root, f"*{base}*grad*cam*layer*.png"),
                os.path.join(root, f"*{base}*Grad*CAM*.png"),
                os.path.join(root, "**", f"{base}_grad_cam_all_classes_layer*.png"),
                os.path.join(root, "**", f"*{base}*grad*cam*.png"),
                os.path.join(root, "**", f"*{base}*Grad*CAM*.png"),
            ]
        )

    found = []

    for pattern in patterns:
        try:
            found.extend(glob.glob(pattern, recursive=True))
        except Exception:
            pass

    found = [
        p for p in found
        if os.path.isfile(p)
        and str(p).lower().endswith((".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"))
    ]

    # Deduplicate while preserving order.
    found = list(dict.fromkeys(found))

    def layer_key(path):
        name = os.path.basename(str(path)).lower()
        match = re.search(r"layer(\d+)", name)
        if match:
            return int(match.group(1))
        return -1

    found = sorted(found, key=layer_key)

    if n_layers is not None and int(n_layers) > 0 and len(found) > int(n_layers):
        found = found[-int(n_layers):]

    return found
